from openpyxl import Workbook


def save_as_excel(l,save_path):
    workbook = Workbook()
    worksheet = workbook.active
    # print(l)
    j = 1
    for k, v in l[0].items():
        worksheet.cell(1, j).value = k
        j += 1

    for i in range(1, len(l) + 1):
        j = 1
        try:
            for k, v in l[i - 1].items():
                worksheet.cell(i + 1, j).value = v
                j += 1
        except:
            # print(l)
            raise
            # continue
    workbook.save(save_path)
